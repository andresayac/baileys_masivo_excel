import requests
import json
import re

import pyqrcode

from PIL import Image
from pyzbar.pyzbar import decode
from io import BytesIO
import base64

import time
import openpyxl


def valid_session(api_url, session_id):

    try:
        url = f"{api_url}/sessions/status/{session_id}"
        response = requests.request("GET", url)
        response_json = json.loads(response.text)
        if not response_json['success']:
            return False
        return True if response_json['data']['status'] in {'authenticated'} else False
    except Exception as e:
        exit("Error al validar la session: " + str(e))

def create_session(api_url, session_id):
    try:
        url = f"{api_url}/sessions/add"

        payload= f"id={session_id}&isLegacy=false"
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        response = requests.request("POST", url, headers=headers, data=payload)
        return json.loads(response.text)


    except Exception as e:
        exit("Error al crear session: " + str(e))
      


def validate_session_name(text):
    return True if re.match(r'^[a-zA-Z0-9]+$', text) else False


def genera_qr(qr_data):
    decode_data = decode(Image.open(BytesIO(base64.b64decode(qr_data))))[0].data
    pyqrcode.create(decode_data).show(quiet_zone=10)

def sleep_check_session(api_url, session_id, total_seconds = 60, interval_seconds = 2):

    time_sleep = total_seconds / interval_seconds
    tmp_time = 0
    while tmp_time < interval_seconds:
        tmp_time += 1
        if valid_session(api_url,session_id) :
            print(f"Session {session_id} creada con Exito.")
            return True
        time.sleep(time_sleep)

    return False  


def send_text_message(api_url, session_id, phone_number, message):
    try:
        url = f"{api_url}/chats/send?id={session_id}"

        payload = json.dumps({
        "receiver": re.escape(phone_number),
        "isGroup": False,
        "message": {
            "text": message
        }
        })
        headers = {
        'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        response_json = json.loads(response.text)
        return True if response_json['data']['status'] in {'authenticated'} else False

    except Exception as e:
        return False


def send_image_message(api_url, session_id, phone_number, message):
    try:
        url = f"{api_url}/chats/send?id={session_id}"

        payload = json.dumps({
        "receiver": re.escape(phone_number),
        "isGroup": False,
        "message": {
            "image": {
            "url": message
            }
        }
        })
        headers = {
        'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        response_json = json.loads(response.text)
        return True if response_json['data']['status'] in {'authenticated'} else False

    except Exception as e:
        return False



def send_video_message(api_url, session_id, phone_number, message):
    try:
        url = f"{api_url}/chats/send?id={session_id}"

        payload = json.dumps({
        "receiver": re.escape(phone_number),
        "isGroup": False,
        "isBase64": False,
        "message": {
            "video": {
            "url": message
            }
        }
        })
        headers = {
        'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        response_json = json.loads(response.text)
        return True if response_json['data']['status'] in {'authenticated'} else False

    except Exception as e:
        return False

def send_gif_message(api_url, session_id, phone_number, message):
    try:
        url = f"{api_url}/chats/send?id={session_id}"

        payload = json.dumps({
        "receiver": re.escape(phone_number),
        "isGroup": False,
        "isBase64": False,
        "message": {
            "video": {
            "url": message,
            "gifPlayback": True
            }
        }
        })
        headers = {
        'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        response_json = json.loads(response.text)
        return True if response_json['data']['status'] in {'authenticated'} else False

    except Exception as e:
        return False


def send_document_message(api_url, session_id, phone_number, message):
    try:
        url = f"{api_url}/chats/send?id={session_id}"

        payload = json.dumps({
        "receiver": re.escape(phone_number),
        "isGroup": False,
        "message": {
            "document": {
            "url": message
            }
        }
        })
        headers = {
        'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        response_json = json.loads(response.text)

    except Exception as e:
        return False


def read_xlsx_file(file_excel):
    workbook = openpyxl.load_workbook(file_excel)
    worksheet = workbook.active
    data = []
    for row in worksheet.iter_rows(min_row=2):
        row_data = []
        for cell in row:
            if cell.value is None:
                continue
            row_data.append(cell.value)
        if len(row_data) == 0:
            continue
        data.append(row_data)
    return data

def create_xlsx_file(data, file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if len(data) == 0:
        exit("No hay datos enviados")
    for row in data:
       sheet.append(row)

    # Guarda el archivo
    workbook.save(file_name)